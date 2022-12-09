import csv
from datetime import datetime
import re
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import openpyxl.utils.cell
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055
}


class Vacancy:
    """Класс, содержащий параметры вакансии

     Attributes:
         name (str): Название вакансии
         salary_from (float): Нижняя граница вилки оклада
         salary_to (float): Верхняя граница вилки оклада
         salary_currency (str): Валюта оклада
         area_name (str): Страна
         published_at (str): Дата публикации
    """
    def __init__(self, items):
        """Инициализирует объект Vacancy, выполняет конвертацию границ оклада в float

        Args:
            items (list): Список значений, формируемых с помощью класса DataSet
        """
        self.name = items[0]
        self.salary_from = float(items[1])
        self.salary_to = float(items[2])
        self.salary_currency = items[3]
        self.area_name = items[4]
        self.published_at = items[5]


class DataSet:
    """Класс, подготавливающий данные из csv-файла для передачи в класс Vacancy

    Attributes:
        file_name (str): Имя файла
        vacancies (list): Список вакансий
    """
    def __init__(self, file_name):
        """Инициализирует объект DataSet

        Args:
            file_name (str): Имя файла
        """
        self.file_name = file_name
        self.vacancies = []

    @staticmethod
    def get_dataset(file_name):
        """Формирует данные

        Args:
            file_name (str): Имя csv-файла

        Returns:
            list: Список вакансий
        """
        data = DataSet.csv_reader(file_name)
        data_array = DataSet.csv_filer(data[0], data[1])
        dataset = DataSet(file_name)
        for item in data_array:
            vacancy_items = [f"{item['name']}", f"{item['salary_from']}", f"{item['salary_to']}",
                             f"{item['salary_currency']}", f"{item['area_name']}", f"{item['published_at']}"]
            vacancy = Vacancy(vacancy_items)
            vacancy.published_at = datetime.strptime(vacancy.published_at, "%Y-%m-%dT%H:%M:%S%z").year
            dataset.vacancies.append(vacancy)
        return dataset

    @staticmethod
    def csv_reader(file_name):
        """Считывает данные из csv-файла

        Args:
            file_name (str): Имя csv-файла

        Returns:
            list: Названия параметров вакансий
            list: Параметры вакансий
        """
        file_csv = open(file_name, encoding="utf_8_sig")
        reader_csv = csv.reader(file_csv)
        list_data = [x for x in reader_csv]
        return list_data[0], [x for x in list_data[1:] if len(x) == len(list_data[0]) and not x.__contains__("")]

    @staticmethod
    def csv_filer(list_naming, reader):
        """Формирует список вакансий

        Args:
            list_naming (list): Названия параметров вакансий
            reader (list): Параматры вакансий

        Returns:
            list: Список всех вакансий с названиями их параметров
        """
        data_vacancies_local = []
        for vacancy in reader:
            data_dictionary = {}
            for title in range(len(vacancy)):
                items = DataSet.remove_html_tags(vacancy[title].split('\n'))
                data_dictionary[list_naming[title]] = items[0] if len(items) == 1 else "; ".join(items)
            data_vacancies_local.append(data_dictionary)
        return data_vacancies_local

    @staticmethod
    def remove_html_tags(vacancy):
        """Удаляет html-теги из параметров вакансии

        Args:
            vacancy (list): Параметры вакансии

        Returns:
            list: Параметры вакансии с удаленными html-тегами
        """
        for title in range(len(vacancy)):
            vacancy[title] = " ".join(re.sub(r"\<[^>]*\>", "", vacancy[title]).split())
        return vacancy


class InputConnect:
    """Класс, отвечающий за сбор статистики по вакансиям. Получает данные от пользователя, передает статистику классу Report

    Attributes:
        file_name (str): Имя файла
        profession_name (str): Название профессии
    """
    def __init__(self):
        """Инициализирует объект InputConnect"""
        self.file_name = input("Введите название файла: ")
        self.profession_name = input("Введите название профессии: ")

    @staticmethod
    def print_data_dict(self, data: DataSet):
        """Выводит на экран статистику о вакансиях

        Args:
            data (DataSet): Список вакансий
        """
        def get_correct_vacancy_rate(data: DataSet):
            """Отвечает за правильный рассчет количества вакансий по городам в процентом отношении к общему количеству вакансий

            Args:
                data (DataSet): Список вакансий

            Returns:
                dict: Отсортированный по убыванию количества вакансий словарь вакансий
            """
            data.vacancy_rate_by_city = {x: round(y / len(data.vacancies), 4) for x, y in
                                         data.vacancy_rate_by_city.items()}
            data.vacancy_rate_by_city = {k: v for k, v in data.vacancy_rate_by_city.items() if math.floor(v * 100 >= 1)}
            return dict(sorted(data.vacancy_rate_by_city.items(), key=lambda item: item[1], reverse=True))
        data.vacancy_rate_by_city = InputConnect.get_vacancy_rate_by_city(data)
        data.salary_by_city = InputConnect.get_salary_by_city(data)
        data.vacancy_rate_by_city = get_correct_vacancy_rate(data)
        data.vacancies_count_by_year = InputConnect.get_vacancies_count_by_name(data, "None")
        data.salary_by_year = InputConnect.get_salary_by_name(data, "None")
        data.vacancies_count_by_profession_name = InputConnect.get_vacancies_count_by_name(data, self.profession_name)
        data.salary_by_profession_name = InputConnect.get_salary_by_name(data, self.profession_name)

        salary_by_year = data.salary_by_year
        vacs_by_years = data.vacancies_count_by_year
        vac_salary_by_years = data.salary_by_profession_name
        vac_counts_by_years = data.vacancies_count_by_profession_name
        salary_by_cities = dict(list(data.salary_by_city.items())[:10])
        vacs_by_cities = dict(list(data.vacancy_rate_by_city.items())[:10])

        Report(salary_by_year, vacs_by_years, vac_salary_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities,
               self.profession_name)

        print(f"Динамика уровня зарплат по годам: ", salary_by_year)
        print(f"Динамика количества вакансий по годам: ", vacs_by_years)
        print(f"Динамика уровня зарплат по годам для выбранной профессии: ", vac_salary_by_years)
        print(f"Динамика количества вакансий по годам для выбранной профессии: ", vac_counts_by_years)
        print(f"Уровень зарплат по городам (в порядке убывания): ", salary_by_cities)
        print(f"Доля вакансий по городам (в порядке убывания): ", vacs_by_cities)

    @staticmethod
    def get_vacancies_count_by_name(data: DataSet, name):
        """Считает количество вакансий по годам

        Args:
            data (DataSet): Список вакансий
            name (str): Название профессии

        Returns:
            dict: Словарь с количеством вакансий по годам
        """
        vacancies_count = {}
        for vacancy in data.vacancies:
            if vacancy.name.__contains__(name) or name == "None":
                InputConnect.set_value_by_name(vacancies_count, vacancy.published_at)
        if len(vacancies_count) == 0:
            return {2022: 0}
        return vacancies_count

    @staticmethod
    def get_salary_by_name(data: DataSet, name):
        """Преобразовывает данные о зарплате у каждой вакансии

        Args:
            data (DataSet): Список вакансий
            name (str): Название профессии

        Returns:
            dict: Словарь с преобразованными данными о зарплатах
        """
        salary_by_name = {}
        for vacancy in data.vacancies:
            if vacancy.name.__contains__(name) or name == "None":
                if not salary_by_name.__contains__(vacancy.published_at):
                    salary_by_name[vacancy.published_at] = InputConnect.convert_currency(vacancy)
                else:
                    salary_by_name[vacancy.published_at] += InputConnect.convert_currency(vacancy)
        if len(salary_by_name) == 0:
            return {2022: 0}
        for key in salary_by_name.keys():
            if name == "None":
                salary_by_name[key] = math.floor(salary_by_name[key] / data.vacancies_count_by_year[key])
            else:
                salary_by_name[key] = math.floor(salary_by_name[key] / data.vacancies_count_by_profession_name[key])
        return salary_by_name

    @staticmethod
    def get_vacancy_rate_by_city(data: DataSet):
        """Приводит статистику вакансий по городам

        Args:
            data (DataSet): Список вакансий

        Returns:
            dict: Словарь со статистикой вакансий по городам
        """
        vacancy_rate = {}
        for vacancy in data.vacancies:
            InputConnect.set_value_by_name(vacancy_rate, vacancy.area_name)
        return vacancy_rate

    @staticmethod
    def set_value_by_name(vacancy_dict: dict, name):
        """Вспомогательный метод для подсчета вакансий по городам

        Args:
            vacancy_dict (dict): Список вакансий
            name (str): Название города
        """
        if not vacancy_dict.__contains__(name):
            vacancy_dict[name] = 1
        else:
            vacancy_dict[name] += 1

    @staticmethod
    def convert_currency(vacancy):
        """Конвертирует валюту в рубли для параметра "Валюта оклада" у вакансий

        Args:
            vacancy (Vacancy): объект класса Vacancy
        """
        rate = currency_to_rub[vacancy.salary_currency]
        return int((vacancy.salary_from * rate + vacancy.salary_to * rate) / 2)

    @staticmethod
    def get_salary_by_city(data: DataSet):
        """Приводит статистику вакансий по уровню зарплат в городах

        Args:
            data (DataSet): Список вакансий

        Returns:
            dict: Отсортированный по убыванию уровня зарплат словарь вакансий
        """
        salary_by_city = {}
        for vacancy in data.vacancies:
            if math.floor(data.vacancy_rate_by_city[vacancy.area_name] / len(data.vacancies) * 100) >= 1:
                if not salary_by_city.__contains__(vacancy.area_name):
                    salary_by_city[vacancy.area_name] = InputConnect.convert_currency(vacancy)
                else:
                    salary_by_city[vacancy.area_name] += InputConnect.convert_currency(vacancy)
        for key in salary_by_city:
            salary_by_city[key] = math.floor(salary_by_city[key] / data.vacancy_rate_by_city[key])
        return dict(sorted(salary_by_city.items(), key=lambda item: item[1], reverse=True))


class Report:
    """Класс, отвечающий за визуализацию статистики вакансий

    Attributes:
        salary_by_year (dict): Уровень зарплат всех вакансий по годам
        vacs_by_years (dict):  Количество всех вакансий по годам
        vac_salary_by_years (dict): Уровень зарплат конкретной профессии по годам
        vac_counts_by_years (dict): Количество вакансий конкретной профессии по годам
        salary_by_cities (dict): Список городов с самыми высокими зарплатами конкретной профессии
        vacs_by_cities (dict): Список с отношениями количества вакансий по конкретной профессии к общему количеству вакансий по городам
        profession (str): Название профессии
    """
    def __init__(self, salary_by_year, vacs_by_years, vac_salary_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities, profession_name):
        """Инициализирует объект Report

        Args:
            salary_by_year (dict): Уровень зарплат всех вакансий по годам
            vacs_by_years (dict):  Количество всех вакансий по годам
            vac_salary_by_years (dict): Уровень зарплат конкретной профессии по годам
            vac_counts_by_years (dict): Количество вакансий конкретной профессии по годам
            salary_by_cities (dict): Список городов с самыми высокими зарплатами конкретной профессии
            vacs_by_cities (dict): Список с отношениями количества вакансий по конкретной профессии к общему количеству вакансий по городам
            profession_name (str): Название профессии
        """
        self.salary_by_year = salary_by_year
        self.vacs_by_years = vacs_by_years
        self.vac_salary_by_years = vac_salary_by_years
        self.vac_counts_by_years = vac_counts_by_years
        self.salary_by_cities = salary_by_cities
        self.vacs_by_cities = vacs_by_cities

        self.profession = profession_name

        Report.generate_excel(self.salary_by_year, self.vacs_by_years, self.vac_salary_by_years, self.vac_counts_by_years,
                              self.salary_by_cities, self.vacs_by_cities, self.profession)

    @staticmethod
    def generate_excel(salary_by_year, vacs_by_years, vac_salary_by_years, vac_counts_by_years, salary_by_cities,
                       vacs_by_cities, profession):
        """Формирует таблицу Excel с данными о вакансиях по выбраннной профессии

        Args:
            salary_by_year (dict): Уровень зарплат всех вакансий по годам
            vacs_by_years (dict):  Количество всех вакансий по годам
            vac_salary_by_years (dict): Уровень зарплат конкретной профессии по годам
            vac_counts_by_years (dict): Количество вакансий конкретной профессии по годам
            salary_by_cities (dict): Список городов с самыми высокими зарплатами конкретной профессии
            vacs_by_cities (dict): Список с отношениями количества вакансий по конкретной профессии к общему количеству вакансий по городам
            profession (str): Название профессии
        """
        wb = Workbook()
        sheet1 = wb.active
        thin = Side(border_style="thin", color="000000")
        sheet1.title = "Статистика по годам"
        sheet2 = wb.create_sheet("Статистика по городам")
        heads1 = ["Год", "Средняя зарплата", f"Средняя зарплата - {profession}",
                  "Количество вакансий", f"Количество вакансий - {profession}"]
        heads2 = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]

        for i, head in enumerate(heads1):
            sheet1.cell(row=1, column=(i + 1), value=head).font = Font(bold=True)
        for i, head in enumerate(heads2):
            sheet2.cell(row=1, column=(i + 1), value=head).font = Font(bold=True)

        for year, value in salary_by_year.items():
            sheet1.append((year, value, vac_salary_by_years[year], vacs_by_years[year], vac_counts_by_years[year]))
        for city, value in salary_by_cities.items():
            sheet2.append({"A": city, "B": value, "C": ""})
        for city, value in vacs_by_cities.items():
            sheet2.append({"D": city, "E": str(value*100)+"%"})

        sheet2.move_range(cell_range="D12:E21", rows=-10)

        for column in sheet1.columns:
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)
        for column in sheet2.columns:
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)

        def as_text(value):
            """Вспомогательный метод, отвечающий за правильную визуализацию ячеек таблицы и конвертацию их в строку

            Args:
                value (int or float to str): Значение ячейки таблицы

            Returns:
                str: Стрковое значение ячейки таблицы
            """
            if value is None:
                return ""
            return str(value)

        for column_cells in sheet1.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            sheet1.column_dimensions[openpyxl.utils.cell.get_column_letter(column_cells[0].column)].width = length + 2
        for column_cells in sheet2.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            sheet2.column_dimensions[openpyxl.utils.cell.get_column_letter(column_cells[0].column)].width = length + 2

        Report.generate_image(salary_by_year, vacs_by_years, vac_salary_by_years, vac_counts_by_years, salary_by_cities,
                       vacs_by_cities, profession)
        wb.save("report_task2.xlsx")

    @staticmethod
    def generate_image(salary_by_year, vacs_by_years, vac_salary_by_years, vac_counts_by_years, salary_by_cities,
                       vacs_by_cities, profession):
        """Формирует изображение с графиками статистики по вакансиям выбраннной профессии

        Args:
            salary_by_year (dict): Уровень зарплат всех вакансий по годам
            vacs_by_years (dict):  Количество всех вакансий по годам
            vac_salary_by_years (dict): Уровень зарплат конкретной профессии по годам
            vac_counts_by_years (dict): Количество вакансий конкретной профессии по годам
            salary_by_cities (dict): Список городов с самыми высокими зарплатами конкретной профессии
            vacs_by_cities (dict): Список с отношениями количества вакансий по конкретной профессии к общему количеству вакансий по городам
            profession (str): Название профессии
        """
        width_coef = 0.4
        other_vacs = 1 - sum([value for value in vacs_by_cities.values()])
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2)

        ax1.set_title("Уровень зарплат по годам")
        bar1 = ax1.bar(np.array(list(salary_by_year.keys())) - 0.4, salary_by_year.values(), width=width_coef)
        bar2 = ax1.bar(np.array(list(salary_by_year.keys())), vac_salary_by_years.values(), width=width_coef)
        ax1.grid(axis="y")
        ax1.set_xticks(np.array(list(salary_by_year.keys())) - 0.2, list(salary_by_year.keys()), rotation=90)
        ax1.xaxis.set_tick_params(labelsize=8)
        ax1.yaxis.set_tick_params(labelsize=8)
        ax1.legend((bar1[0], bar2[0]), ("средняя з/п", "з/п " + profession.lower()), prop={"size": 8})

        ax2.set_title("Количество вакансий по годам", fontdict={'fontsize': 11})
        bar1 = ax2.bar(np.array(list(vacs_by_years.keys())) - 0.4, vacs_by_years.values(), width=width_coef)
        bar2 = ax2.bar(np.array(list(vacs_by_years.keys())), vac_counts_by_years.values(), width=width_coef)
        ax2.set_xticks(np.array(list(vacs_by_years.keys())) - 0.2, list(vacs_by_years.keys()), rotation=90)
        ax2.grid(axis="y")
        ax2.xaxis.set_tick_params(labelsize=8)
        ax2.yaxis.set_tick_params(labelsize=8)
        ax2.legend((bar1[0], bar2[0]), ("Количество вакансий", "Количество вакансий\n" + profession.lower()),
                   prop={"size": 8})

        ax3.set_title("Уровень зарплат по городам")
        ax3.barh(list([str(a).replace(" ", "\n").replace("-", "-\n") for a in reversed(list(salary_by_cities.keys()))]),
                 list(reversed(list(salary_by_cities.values()))), color="blue", height=0.5, align="center")
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis="x")

        ax4.set_title("Доля вакансий по городам")
        ax4.pie(list(vacs_by_cities.values()) + [other_vacs], labels=list(vacs_by_cities.keys()) + ["Другие"], textprops={"fontsize": 6})

        plt.tight_layout()
        plt.savefig("graph_task2.png")


input_data = InputConnect()
data = DataSet.get_dataset(input_data.file_name)
input_data.print_data_dict(input_data, data)
