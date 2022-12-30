from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit
import pandas as pd

# currency_to_rub = {
#     "AZN": 35.68,
#     "BYR": 23.91,
#     "EUR": 59.90,
#     "GEL": 21.74,
#     "KGS": 0.76,
#     "KZT": 0.13,
#     "RUR": 1,
#     "UAH": 1.64,
#     "USD": 60.66,
#     "UZS": 0.0055
# }


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.salary_by_year = {}
        self.vacancies_count_by_year = {}
        self.salary_by_profession = {}
        self.vacancies_count_by_profession = {}
        self.dict_lict = []


class InputConnect:
    def __init__(self):
        self.file_name, self.profession_name = InputConnect.get_parameters()

    @staticmethod
    def get_parameters():
        file_name = input("Введите название файла: ") # converted_dataframe.csv
        profession_name = input("Введите название профессии: ") # Программист
        return file_name, profession_name

    @staticmethod
    def print_data(self, data: DataSet):
        df = pd.read_csv(data.file_name)
        df['salary'] = df['salary'].fillna(0)
        df['salary'] = df['salary'].astype("int64")
        df["published_at"] = df["published_at"].apply(lambda x: int(x[:4]))
        years = df["published_at"].unique()
        df_vacancy = df["name"].str.contains(self.profession_name)
        for year in years:
            filter_by_year = df["published_at"] == year
            data.salary_by_year[year] = int(df[filter_by_year]["salary"].mean())
            data.vacancies_count_by_year[year] = len(df[filter_by_year])
            data.salary_by_profession[year] = int(df[df_vacancy & filter_by_year]["salary"].mean())
            data.vacancies_count_by_profession[year] = len(df[df_vacancy & filter_by_year])


class Report:
    @staticmethod
    def generate_excel(profession_name, data: DataSet):
        wb = Workbook()
        sheet1 = wb.worksheets[0]
        sheet1.title = "Статистика по годам"
        heads = ["Год", "Средняя зарплата", f"Средняя зарплата - {profession_name}",
                   "Количество вакансий", f"Количество вакансий - {profession_name}"]

        def as_text(value):
            if value is None:
                return ""
            return str(value)

        for i, cell in enumerate(sheet1['A1':'E1'][0]):
            cell.value = heads[i]
            cell.font = Font(size=11, b=True)
        for key in data.salary_by_year:
            sheet1.append([key, data.salary_by_year[key], data.salary_by_profession[key],
                           data.vacancies_count_by_year[key], data.vacancies_count_by_profession[key]])
        for column_cells in sheet1.columns:
            for cell in column_cells:
                bd = Side(style="thin", color="000000")
                cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        for column_cells in sheet1.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            sheet1.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        wb.save("report.xlsx")
        return

    @staticmethod
    def generate_image(profession_name, data: DataSet):
        width = 0.3
        nums = np.arange(len(data.salary_by_year.keys()))
        dx1 = nums - width / 2
        dx2 = nums + width / 2
        fig = plt.figure()
        ax = fig.add_subplot(211)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(dx1, data.salary_by_year.values(), width, label="средняя з/п")
        ax.bar(dx2, data.salary_by_profession.values(), width, label=f"з/п {profession_name.lower()}")
        ax.set_xticks(nums, data.salary_by_year.keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')
        ax = fig.add_subplot(212)
        ax.set_title("Количество вакансии по годам")
        ax.bar(dx1, data.vacancies_count_by_year.values(), width, label="Количество вакансии")
        ax.bar(dx2, data.vacancies_count_by_profession.values(), width,
               label=f"Количество вакансии\n{profession_name.lower()}")
        ax.set_xticks(nums, data.salary_by_year.keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')
        plt.tight_layout()
        fig.set_size_inches(9.5, 7.5)
        plt.savefig("graph1.png", dpi=120)
        return

    @staticmethod
    def generate_pdf(profession, data: DataSet):
        Report.generate_excel(profession, data)
        Report.generate_image(profession, data)
        book = load_workbook("report.xlsx")
        sheet1 = book.active
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template_1.html")
        pdf_template = template.render({'name': profession, 'sheet1': sheet1})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltox\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


input_data = InputConnect()
data = DataSet(input_data.file_name)
InputConnect.print_data(input_data, data)
Report.generate_pdf(input_data.profession_name, data)