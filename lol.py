import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl.reader.excel import load_workbook
import pandas as pd


class DataSet:

    def __init__(self, file_name):
        self.file_name = file_name
        self.salary_by_year = dict()
        self.vacancies_count_by_year = dict()
        self.salary_by_profession_name = dict()
        self.vacancies_count_by_profession_name = dict()
        self.salary_by_city = dict()
        self.vacancy_rate_by_city = dict()
        self.dict_lict = list()


class InputConnect:
    def __init__(self):
        self.file_name, self.profession_name, self.area_name = InputConnect.get_params()

    @staticmethod
    def get_params():
        file_name = input("Введите название файла: ")
        profession_name = input("Введите название профессии: ")
        area_name = input("Введите название региона: ")
        return file_name, profession_name, area_name

    @staticmethod
    def print_data_dict(self, data: DataSet):
        df = pd.read_csv(data.file_name)
        df['salary'] = df['salary'].fillna(0)
        df['salary'] = df['salary'].astype("int64")
        df["published_at"] = df["published_at"].apply(lambda d: int(d[:4]))
        years = df["published_at"].unique()
        df_vacancy = df["name"].str.contains(self.profession_name)
        df_area = df["area_name"].str.contains(self.area_name)

        for year in years:
            filter_by_year = df["published_at"] == year
            data.salary_by_year[year] = int(df[filter_by_year]["salary"].mean())
            data.vacancies_count_by_year[year] = len(df[filter_by_year])
            data.salary_by_profession_name[year] = int(df[df_vacancy & filter_by_year & df_area]["salary"].mean())
            data.vacancies_count_by_profession_name[year] = len(df[df_vacancy & filter_by_year & df_area])

        count = len(df)
        df["count"] = df.groupby("area_name")["area_name"].transform("count")
        df_norm = df[df["count"] > 0.01 * count]
        df_area = df_norm.groupby("area_name", as_index=False)["salary"].mean().sort_values(by="salary", ascending=False)
        df_area["salary"] = df_area["salary"].apply(lambda x: int(x))
        df_area10 = df_area.head(10)
        data.salary_by_city = dict(zip(df_area10["area_name"], df_area10["salary"]))

        data.vacancy_rate_by_city = {k: round(v / count, 4) for k, v in dict(df["area_name"].value_counts()).items()}


class Report:
    @staticmethod
    def generate_excel(profession_name, data: DataSet):
        def as_text(value):
            if value is None:
                return ""
            return str(value)

        def set_max_length(worksheet):
            for column_cells in worksheet.columns:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        def set_format_percent(worksheet):
            for i, column_cells in enumerate(worksheet.columns):
                if i == 4:
                    for cell in column_cells:
                        cell.number_format = FORMAT_PERCENTAGE_00

        def set_border_style(worksheet):
            for column_cells in worksheet.columns:
                for cell in column_cells:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        def set_headers(headers, head_range):
            for i, cell in enumerate(head_range):
                cell.value = headers[i]
                cell.font = Font(size=11, b=True)

        wb = Workbook()
        sheet_1 = wb.worksheets[0]
        sheet_1.title = "Статистика по годам"
        sheet_2 = wb.create_sheet("Статистика по городам")
        headers = ["Год", "Средняя зарплата", f"Средняя зарплата - {profession_name}",
                   "Количество вакансий", f"Количество вакансий - {profession_name}"]
        set_headers(headers, sheet_1['A1':'E1'][0])

        for key in data.salary_by_year:
            sheet_1.append([key, data.salary_by_year[key], data.salary_by_profession_name[key],
                            data.vacancies_count_by_year[key], data.vacancies_count_by_profession_name[key]])
        set_border_style(sheet_1)
        set_max_length(sheet_1)

        set_headers(["Город", "Уровень зарплат"], sheet_2['A1':'B1'][0])
        set_headers(["Город", "Доля вакансий"], sheet_2['D1':'E1'][0])
        sheet_2.column_dimensions['C'].width = 2
        city_keys = list(data.vacancy_rate_by_city.keys())
        for i, key in enumerate(data.salary_by_city.keys()):
            sheet_2.append([key, data.salary_by_city[key], None, city_keys[i], data.vacancy_rate_by_city[city_keys[i]]])

        for i, column_cells in enumerate(sheet_2.columns):
            for cell in column_cells:
                if i != 2:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        set_format_percent(sheet_2)
        set_max_length(sheet_2)
        wb.save("report.xlsx")
        return

    @staticmethod
    def generate_image(profession_name, data: DataSet):
        def myfunc(item):
            if item.__contains__(' '):
                return item[:item.index(' ')] + '\n' + item[item.index(' ') + 1:]
            elif item.__contains__('-'):
                return item[:item.index('-')] + '-\n' + item[item.index('-') + 1:]
            return item

        width = 0.3
        nums = np.arange(len(data.salary_by_year.keys()))
        dx1 = nums - width / 2
        dx2 = nums + width / 2

        fig = plt.figure()
        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(dx1, data.salary_by_year.values(), width, label="средняя з/п")
        ax.bar(dx2, data.salary_by_profession_name.values(), width, label=f"з/п {profession_name.lower()}")
        ax.set_xticks(nums, data.salary_by_year.keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансии по годам")
        ax.bar(dx1, data.vacancies_count_by_year.values(), width, label="Количество вакансии")
        ax.bar(dx2, data.vacancies_count_by_profession_name.values(), width,
               label=f"Количество вакансии\n{profession_name.lower()}")
        ax.set_xticks(nums, data.salary_by_year.keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        cities = list(map(myfunc, tuple(data.salary_by_city.keys())))
        y_pos = np.arange(len(cities))
        ax.barh(y_pos, list(data.salary_by_city.values()), align='center')
        ax.set_yticks(y_pos, labels=cities)
        ax.invert_yaxis()
        ax.grid(True, axis='x')

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансии по годам")
        labels = list(dict(list(data.vacancy_rate_by_city.items())[:10]).keys())
        labels.insert(0, "Другие")
        vals = list(dict(list(data.vacancy_rate_by_city.items())[:10]).values())
        vals.insert(0, 1 - sum(list(dict(list(data.vacancy_rate_by_city.items())[:10]).values())))
        ax.pie(vals, labels=labels, startangle=0, textprops={"fontsize": 6})
        plt.tight_layout()
        fig.set_size_inches(9.5, 7.5)
        plt.savefig("graph.png", dpi=120)
        return

    @staticmethod
    def generate_pdf(profession_name, data: DataSet):
        Report.generate_excel(profession_name, data)
        Report.generate_image(profession_name, data)
        name = profession_name
        image_file = "graph.png"
        book = load_workbook("report.xlsx")
        sheet_1 = book.active
        sheet_2 = book['Статистика по городам']
        for row in range(2, sheet_2.max_row + 1):
            for col in range(4, 6):
                if type(sheet_2.cell(row, col).value).__name__ == "float":
                    sheet_2.cell(row, col).value = str(round(sheet_2.cell(row, col).value * 100, 2)) + '%'

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template_2.html")
        pdf_template = template.render({'name': name, 'image_file': image_file, 'sheet_1': sheet_1, 'sheet_2': sheet_2})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltox\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


inputparam = InputConnect()
dataset = DataSet(inputparam.file_name)
InputConnect.print_data_dict(inputparam, dataset)
Report.generate_pdf(inputparam.profession_name, dataset)